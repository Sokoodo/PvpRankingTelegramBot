import telebot
from openpyxl import load_workbook
from telebot.types import InlineKeyboardMarkup, InlineKeyboardButton

API_TOKEN = ''
bot = telebot.TeleBot(API_TOKEN)
leagues = ['GreatLeague', 'UltraLeague', 'MasterLeague']


def fill_names_list():
    workbook = load_workbook(filename='pokemon_leagues.xlsx', read_only=True)
    names_sheet = workbook['pokemon_names']
    names = []
    for row in names_sheet.rows:
        for cell in row:
            if cell.value is not None:
                cell_val = '' + cell.value
                names.append(cell_val.lower())
    return names


pokemon_names = fill_names_list()


@bot.message_handler(commands=['top20'])
def send_top20(message):
    bot.send_message(message.chat.id, "Select the league", reply_markup=gen_leagues())


def gen_leagues():
    markup = InlineKeyboardMarkup()
    markup.row_width = 1
    for i in range(0, 3):
        markup.add(InlineKeyboardButton(leagues[i], callback_data=leagues[i]))
    return markup


@bot.callback_query_handler(func=lambda query: query.data in leagues)
def callback_query(call):
    workbook = load_workbook(filename='pokemon_leagues.xlsx', read_only=True)
    if call.data == 'GreatLeague':
        league_sheet = workbook['great_league']
        top20_str = 'TOP 20 GreatLeague:\n'
    elif call.data == 'UltraLeague':
        league_sheet = workbook['ultra_league']
        top20_str = 'TOP 20 UltraLeague:\n'
    else:
        league_sheet = workbook['master_league']
        top20_str = 'TOP 20 MasterLeague:\n'
    i = 0
    for row in league_sheet.rows:
        if i < 20:
            for cell in row:
                if cell.value is not None and cell.value != 'Pokemon' and cell.value != 'Score':
                    if cell.column_letter == 'A':
                        i += 1
                        top20_str = top20_str + str(i) + '- ' + cell.value
                    if cell.column_letter == 'B':
                        top20_str = top20_str + ' ' + str(cell.value) + '\n'
        else:
            bot.send_message(call.from_user.id, top20_str)
            break


@bot.message_handler(commands=['help', 'start'])
def send_welcome(message):
    bot.reply_to(message, """\
Hi there, I am the PvpRanking Bot.
You can start with commands /top20\
""")


@bot.message_handler(func=lambda message: message.text.lower() in pokemon_names)
def find_pokemon_info(message):
    message_name = message.text.lower()
    pokemon_info = ''
    for name in pokemon_names:
        if message_name == name:
            pokemon_info = get_full_info(name)
    for info in pokemon_info:
        bot.send_message(message.chat.id, info, parse_mode='Markdown')


def get_full_info(name):
    great_info = get_info_leagues(name, 'great_league')
    ultra_info = get_info_leagues(name, 'ultra_league')
    master_info = get_info_leagues(name, 'master_league')
    info_concat = []
    if great_info != '':
        info_concat.append(great_info)
    if ultra_info != '':
        info_concat.append(ultra_info)
    if master_info != '':
        info_concat.append(master_info)
    return info_concat


def get_info_leagues(name, str_league):
    wb = load_workbook(filename='pokemon_leagues.xlsx', read_only=True)
    league_sheet = wb[str_league]
    for row in league_sheet.rows:
        if row[0].value == name:
            for cell in row:
                cell_row = cell.row
                cell_col = cell.column
                score = '' + str(league_sheet.cell(cell_row, cell_col + 1).value)
                position = '' + str(league_sheet.cell(cell_row, cell_col + 7).value)
                fast_move = '' + str(league_sheet.cell(cell_row, cell_col + 4).value)
                charged_move1 = '' + str(league_sheet.cell(cell_row, cell_col + 5).value)
                charged_move2 = '' + str(league_sheet.cell(cell_row, cell_col + 6).value)
                info_string = '*' + str_league + ':* \n\n*Position* \n' + position + '\n\n*Score* \n' + score + '\n\n*FastMove* \n' + fast_move + '\n\n*ChargedMoves* \n' + charged_move1 + ' \n' + charged_move2
                return info_string
    return ''


@bot.message_handler(func=lambda message: message.text.lower() not in pokemon_names)
def echo_message(message):
    bot.reply_to(message, 'I don\'t know what to do, try with /help or /top20.')


bot.infinity_polling()
